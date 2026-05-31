"""
重新导入所有 Excel 数据到 buyer_mapping 表（包含日期、货物名称、成本、快递单号）
"""
import openpyxl
import json
import urllib.request
import sys
import time

SUPABASE_URL = "https://jlzsonjjfgojmwgghxbl.supabase.co"
SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImpsenNvbmpqZmdvam13Z2doeGJsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjkxNjY4OSwiZXhwIjoyMDkyNDkyNjg5fQ.qTAqHLBeWyVUfV8uxdP2-55EFI7kyh4aJ2RGJHrhQTo"

def api_request(method, path, body=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    data = json.dumps(body, ensure_ascii=False).encode() if body else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as e:
        raw = e.read()
        try:
            err = json.loads(raw)
            print(f"  API Error: {err.get('message', err)}")
        except:
            print(f"  API Error: HTTP {e.code}")
        return None
    except Exception as e:
        print(f"  Network Error: {e}")
        return None

def normalize_date(val):
    """将日期值标准化为字符串"""
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        # Excel 序列号日期
        if val > 40000 and val < 60000:
            from datetime import datetime, timedelta
            try:
                dt = datetime(1899, 12, 30) + timedelta(days=int(val))
                return dt.strftime('%Y-%m-%d')
            except:
                pass
        return str(val)[:10]
    s = str(val).strip()
    # 处理中文日期格式 2026.5.1 -> 2026-05-01
    if '.' in s and len(s) <= 12:
        parts = s.split('.')
        try:
            parts = [str(int(p)) for p in parts if p]
            if len(parts) == 3:
                return f'{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}'
        except:
            pass
    return s[:20] if s else ''

def parse_old_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} (rows: {ws.max_row}, cols: {ws.max_column}) ---")

        # 确定标题行
        header_row = 1
        first_cell = str(ws.cell(1, 1).value or '').strip()
        if '明细' in first_cell or '使用' in first_cell or first_cell.startswith('1'):
            header_row = 2

        headers = {}
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(header_row, c).value or '').strip()
            if val:
                headers[c] = val

        # 列映射
        order_col = None
        account_col = None
        track_col = None
        date_col = None     # A列 = 发货时间
        name_col = None     # B列 = 发货名称
        cost_col = None     # E列 = 成本

        for col, h in headers.items():
            if '订单号' in h or '订单编号' in h:
                order_col = col
            if '购买' in h and ('账号' in h or '人' in h):
                account_col = col
            if ('快递' in h or '运单' in h or '物流' in h) and '单号' in h:
                track_col = col
            if col == 1 and ('发货时间' in h or '发货日期' in h or '日期' in h or '时间' in h):
                date_col = 1
            if col == 2 and ('名称' in h or '商品' in h or '产品' in h):
                name_col = 2
            if '成本' in h:
                cost_col = col

        # Sheet1 特殊处理
        if sheet_name == 'Sheet1':
            order_col, track_col, account_col = 1, 3, 5
            date_col, name_col, cost_col = None, None, None

        # SKU列 = C列 (3)，但9月/10月的C列是订单号，不是SKU
        sku_col = None
        if sheet_name not in ['9月对账用账单', '10月对账用账单', 'Sheet1']:
            sku_col = 3  # C列

        # 默认位置（A=日期, B=名称）
        if not date_col:
            date_col = 1
        if not name_col:
            name_col = 2
        # 成本列按sheet区分
        if not cost_col:
            if sheet_name in ['9月对账用账单', '10月对账用账单']:
                cost_col = 4  # D列
            else:
                cost_col = 5  # E列

        # 快递单号的自动检测（跳过与订单号相同的列）
        if not track_col:
            for candidate_col in [8, 7]:
                if candidate_col != order_col and ws.max_column >= candidate_col:
                    sample = str(ws.cell(header_row + 1, candidate_col).value or '').strip()
                    if sample and (sample.upper().startswith(('SF', 'YT', 'JT', 'JD', 'ST')) or (len(sample) > 10 and not sample.startswith('-'))):
                        track_col = candidate_col
                        break

        print(f"  订单号={order_col}, 日期={date_col}, 名称={name_col}, 成本={cost_col}, 账号={account_col}, 快递={track_col}, SKU={sku_col}")

        if not order_col:
            print(f"  SKIP: 找不到订单号列")
            continue

        sheet_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            order_no = str(ws.cell(r, order_col).value or '').strip()
            if not order_no or order_no == 'None' or len(order_no) < 5:
                continue

            def safe_str(rc):
                v = ws.cell(r, rc).value
                return str(v).strip() if v is not None and str(v).strip() != 'None' else ''

            records.append({
                'order_no': order_no,
                'order_date': normalize_date(ws.cell(r, date_col).value) if date_col else '',
                'product_name': safe_str(name_col) if name_col else '',
                'cost': safe_str(cost_col) if cost_col else '',
                'sku': safe_str(sku_col) if sku_col else '',
                'account': safe_str(account_col) if account_col else '',
                'tracking_no': safe_str(track_col) if track_col else '',
                'source': f'旧表/{sheet_name}'
            })
            sheet_count += 1

        print(f"  导入: {sheet_count} 条")

    return records

def parse_new_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} (rows: {ws.max_row}) ---")

        # A=发货时间, B=发货名称, C=SKU, D=订单号, E=成本, F=购买账号, H=快递单号
        sheet_count = 0
        for r in range(2, ws.max_row + 1):
            order_no = str(ws.cell(r, 4).value or '').strip()
            if not order_no or order_no == 'None' or len(order_no) < 5:
                continue

            def safe_str(rc):
                v = ws.cell(r, rc).value
                return str(v).strip() if v is not None and str(v).strip() != 'None' else ''

            records.append({
                'order_no': order_no,
                'order_date': normalize_date(ws.cell(r, 1).value),
                'product_name': safe_str(2),
                'cost': safe_str(5),
                'sku': safe_str(3),
                'account': safe_str(6),
                'tracking_no': safe_str(8),
                'source': f'新表/{sheet_name}'
            })
            sheet_count += 1

        print(f"  导入: {sheet_count} 条")

    return records

def main():
    all_records = []

    # 解析旧表
    old_path = r'C:\Users\17580\Desktop\账单存放\三月份以前对账表.xlsx'
    print("=" * 50)
    print("解析三月份以前对账表...")
    old_records = parse_old_excel(old_path)
    all_records.extend(old_records)
    print(f"旧表总计: {len(old_records)} 条")

    # 解析新表
    new_path = r'C:\Users\17580\Desktop\4月份对账表.xlsx'
    print("\n" + "=" * 50)
    print("解析4月份对账表...")
    new_records = parse_new_excel(new_path)
    all_records.extend(new_records)
    print(f"新表总计: {len(new_records)} 条")

    # 统计
    with_date = sum(1 for r in all_records if r['order_date'])
    with_name = sum(1 for r in all_records if r['product_name'])
    with_cost = sum(1 for r in all_records if r['cost'])
    with_account = sum(1 for r in all_records if r['account'])
    with_tracking = sum(1 for r in all_records if r['tracking_no'])
    print(f"\n{'='*50}")
    print(f"总计: {len(all_records)} 条")
    print(f"含日期: {with_date}, 含名称: {with_name}, 含成本: {with_cost}")
    print(f"含账号: {with_account}, 含快递: {with_tracking}")

    # 去重合并
    seen = {}
    for r in all_records:
        key = r['order_no']
        if key in seen:
            prev = seen[key]
            for field in ['order_date', 'product_name', 'cost', 'sku', 'account', 'tracking_no']:
                if r[field] and not prev[field]:
                    prev[field] = r[field]
            prev['source'] += '; ' + r['source']
        else:
            seen[key] = r
    deduped = list(seen.values())
    print(f"去重后: {len(deduped)} 条")

    # 保存样本
    output_path = r'C:\Users\17580\Desktop\return-app\import_data.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(deduped[:10], f, ensure_ascii=False, indent=2)

    # 导入到 Supabase
    print(f"\n清空 buyer_mapping...")
    api_request('DELETE', 'buyer_mapping?id=gt.0')

    batch_size = 200
    total = len(deduped)
    imported = 0
    errors = 0

    print(f"导入 {total} 条...")
    for i in range(0, total, batch_size):
        batch = deduped[i:i + batch_size]
        clean_batch = []
        for r in batch:
            clean_batch.append({
                'order_no': str(r.get('order_no', ''))[:255],
                'order_date': str(r.get('order_date', ''))[:50],
                'product_name': str(r.get('product_name', ''))[:255],
                'cost': str(r.get('cost', ''))[:50],
                'sku': str(r.get('sku', ''))[:50],
                'account': str(r.get('account', ''))[:255],
                'tracking_no': str(r.get('tracking_no', ''))[:255],
                'source': str(r.get('source', ''))[:255]
            })

        time.sleep(0.1)  # 避免频率限制
        if api_request('POST', 'buyer_mapping', clean_batch) is not None:
            imported += len(batch)
            pct = imported * 100 // total
            print(f"  {imported}/{total} ({pct}%)")
        else:
            errors += 1
            if errors <= 3:
                print(f"  重试...")
                time.sleep(1)
                if api_request('POST', 'buyer_mapping', clean_batch) is not None:
                    imported += len(batch)
                    continue
            print(f"  批次失败，停止")
            break

    if errors == 0:
        print(f"\n完成！{imported} 条记录")
    else:
        print(f"\n部分完成：{imported}/{total}")

if __name__ == '__main__':
    main()
